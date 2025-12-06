"""
单文件工具（PyQt5）——可自由配置要操作的属性/列名映射。
新增功能：满红导出（在原 sheet 中插入新行作为满红条目）。
修改字段：编辑下方 CONFIG 区的 ATTRS 与 ATTR_COL_MAP。
依赖：pandas openpyxl PyQt5
pip install pandas openpyxl PyQt5
"""
import os
import sys
import site
from typing import Optional, Union, Dict
import pandas as pd

def find_and_set_qt_plugins():
    """
    尝试定位包含 qwindows* 的 platforms 目录并把它设置到 QT_QPA_PLATFORM_PLUGIN_PATH。
    必须在导入 PyQt5.QtWidgets 之前运行。
    """
    candidates = []
    # PyQt5 模块自带位置
    try:
        import PyQt5
        candidates.append(os.path.join(os.path.dirname(PyQt5.__file__), "Qt", "plugins", "platforms"))
    except Exception:
        pass
    # 常见 site-packages 推测位置
    try:
        for p in set(site.getsitepackages()):
            candidates.append(os.path.join(p, "PyQt5", "Qt", "plugins", "platforms"))
            candidates.append(os.path.join(p, "PyQt5_plugins", "platforms"))
    except Exception:
        pass
    # sys.prefix / conda 常见位置
    candidates.append(os.path.join(sys.prefix, "Lib", "site-packages", "PyQt5", "Qt", "plugins", "platforms"))
    candidates.append(os.path.join(sys.prefix, "Library", "plugins", "platforms"))
    if os.environ.get("CONDA_PREFIX"):
        cp = os.environ["CONDA_PREFIX"]
        candidates.append(os.path.join(cp, "Library", "plugins", "platforms"))
        candidates.append(os.path.join(cp, "plugins", "platforms"))

    # 检查候选目录
    for c in candidates:
        if c and os.path.isdir(c):
            for f in os.listdir(c):
                if f.lower().startswith("qwindows"):
                    os.environ["QT_QPA_PLATFORM_PLUGIN_PATH"] = c
                    print("已设置 QT_QPA_PLATFORM_PLUGIN_PATH =", c)
                    return c

    # 回退：在 site-packages 目录中递归查找 qwindows*
    roots = set(site.getsitepackages() if hasattr(site, "getsitepackages") else [sys.prefix])
    for root in roots:
        for dirpath, dirnames, filenames in os.walk(root):
            for f in filenames:
                if f.lower().startswith("qwindows"):
                    os.environ["QT_QPA_PLATFORM_PLUGIN_PATH"] = dirpath
                    print("在路径中找到 qwindows 插件，已设置 QT_QPA_PLATFORM_PLUGIN_PATH =", dirpath)
                    return dirpath

    print("未找到 Qt platforms/qwindows 插件目录。请重装 PyQt5 或手动设置 QT_QPA_PLATFORM_PLUGIN_PATH。")
    return None

_find = find_and_set_qt_plugins()

from PyQt5.QtCore import Qt, QAbstractTableModel, QModelIndex, QVariant
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QFileDialog, QLineEdit, QLabel, QTableView, QMessageBox, QFormLayout,
    QCheckBox, QDialog, QDialogButtonBox, QSpinBox
)
from PyQt5.QtGui import QColor, QBrush

# ---------------- CONFIG: 在这里修改要操作的字段与映射 ----------------
# 逻辑名称（界面显示使用）
ATTRS = ["武力", "智力", "政治", "魅力", "防御", "速度"]

# 映射到 Excel 表头（header=1 时使用第二行作为列名）。
# 每个键为逻辑名，value/init/growth 指向 Excel 中实际列名。
ATTR_COL_MAP = {
    "武力": {"value": "武力", "init": "武力初始", "growth": "武力成长"},
    "智力": {"value": "智力", "init": "智力初始", "growth": "智力成长"},
    "政治": {"value": "政治", "init": "政治初始", "growth": "政治成长"},
    "魅力": {"value": "魅力", "init": "魅力初始", "growth": "魅力成长"},
    "防御": {"value": "防御", "init": "防御初始", "growth": "防御成长"},
    "速度": {"value": "速度", "init": "速度初始", "growth": "速度成长"},
}

# 全局后备后缀（如果 ATTR_COL_MAP 未提供某项，则使用后缀策略）
INIT_SUFFIX = "初始"
GROWTH_SUFFIX = "成长"
DEFAULT_ADD_VALUE = 50
GROWTH_MULT = 49
# -------------------------------------------------------------------

class ExcelHandler:
    def __init__(self):
        self.df: Optional[pd.DataFrame] = None
        self.path: Optional[str] = None
        self.sheet_name: str = "hero"

    def _col_names_for(self, attr: str):
        m = ATTR_COL_MAP.get(attr, {})
        value_col = m.get("value", attr)
        init_col = m.get("init", f"{attr}{INIT_SUFFIX}")
        growth_col = m.get("growth", f"{attr}{GROWTH_SUFFIX}")
        return value_col, init_col, growth_col

    def load(self, path: str, sheet_name: str = "hero") -> pd.DataFrame:
        self.path = path
        self.sheet_name = sheet_name
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, header=1, engine="openpyxl")
        except Exception as e:
            raise RuntimeError(f"读取 Excel 失败: {e}")
        self.df = df.copy()
        self._ensure_required_columns()
        self._compute_base_and_add()
        return self.df

    def _ensure_required_columns(self):
        if self.df is None:
            raise RuntimeError("数据未加载")
        needed = ["ID", "name"]
        for a in ATTRS:
            v, init, grow = self._col_names_for(a)
            needed += [v, init, grow]
        needed = list(dict.fromkeys(needed))
        missing = [c for c in needed if c not in self.df.columns]
        if missing:
            raise RuntimeError(f"缺少必要列: {missing}\n请在文件顶部的 ATTR_COL_MAP 中为对应属性指定实际列名（区分大小写）。")

    def _compute_base_and_add(self):
        if self.df is None:
            return
        df = self.df
        for a in ATTRS:
            value_col, init_col, growth_col = self._col_names_for(a)
            base_col = f"base_{a}"
            add_col = f"add_{a}"
            # 安全读取列，缺失时填 0
            init_series = df[init_col].fillna(0).astype(float) if init_col in df.columns else 0.0
            growth_series = df[growth_col].fillna(0).astype(float) if growth_col in df.columns else 0.0
            value_series = df[value_col].fillna(0).astype(float) if value_col in df.columns else 0.0
            df[base_col] = init_series + growth_series * GROWTH_MULT
            # 先保留文件已有 add_ 列（若存在），否则用 value-base 计算；最终强制为 int
            if add_col in df.columns:
                df[add_col] = pd.to_numeric(df[add_col], errors="coerce").fillna(0).round(0).astype(int)
            else:
                add_series = (value_series - df[base_col]).round(0).fillna(0).astype(int)
                df[add_col] = add_series

        # 计算加点和（整数）
        add_cols = [f"add_{a}" for a in ATTRS if f"add_{a}" in df.columns]
        if add_cols:
            df["add_sum"] = df[add_cols].sum(axis=1).fillna(0).astype(int)
        else:
            df["add_sum"] = 0

        # 默认目标判断（保持原逻辑）
        base_group = [f"base_{x}" for x in ["武力", "智力", "防御", "速度"] if f"base_{x}" in df.columns]
        if base_group:
            df["_default_target"] = df[base_group].idxmax(axis=1)
            df["_default_target_attr"] = df["_default_target"].str.replace(r"^base_", "", regex=True)
        else:
            df["_default_target"] = None
            df["_default_target_attr"] = None

        def is_default(row):
            target = row.get("_default_target_attr")
            for a in ATTRS:
                expected = int(DEFAULT_ADD_VALUE) if a == target else 0
                if int(round(float(row.get(f"add_{a}", 0)))) != expected:
                    return False
            return True
        df["is_default_add"] = df.apply(is_default, axis=1)
        self.df = df

    def get_hero(self, key: Union[int, str]) -> Optional[pd.Series]:
        if self.df is None:
            return None
        mask_id = self.df["ID"].astype(str) == str(key)
        if mask_id.any():
            return self.df[mask_id].iloc[0]
        mask_name = self.df["name"].astype(str).str.contains(str(key), na=False)
        if mask_name.any():
            return self.df[mask_name].iloc[0]
        return None

    def search(self, term: str) -> pd.DataFrame:
        if self.df is None:
            return pd.DataFrame()
        t = str(term)
        mask = self.df["name"].astype(str).str.contains(t, na=False) | (self.df["ID"].astype(str) == t)
        return self.df[mask].copy()

    def update_add_points(self, id_or_name: Union[int, str], new_adds: Dict[str, float]):
        if self.df is None:
            raise RuntimeError("数据未加载")
        idx = None
        mask_id = self.df["ID"].astype(str) == str(id_or_name)
        if mask_id.any():
            idx = self.df[mask_id].index[0]
        else:
            mask_name = self.df["name"].astype(str).str.contains(str(id_or_name), na=False)
            if mask_name.any():
                idx = self.df[mask_name].index[0]
        if idx is None:
            raise KeyError("未找到指定英雄")
        for a, v in new_adds.items():
            if a not in ATTRS:
                continue
            add_int = int(round(v))
            # 只把 add_* 存为整数
            self.df.at[idx, f"add_{a}"] = add_int
            value_col, init_col, growth_col = self._col_names_for(a)
            if f"base_{a}" not in self.df.columns:
                init_val = float(self.df.at[idx, init_col]) if init_col in self.df.columns else 0.0
                growth_val = float(self.df.at[idx, growth_col]) if growth_col in self.df.columns else 0.0
                self.df.at[idx, f"base_{a}"] = init_val + growth_val * GROWTH_MULT
            try:
                base_val = float(self.df.at[idx, f"base_{a}"])
            except Exception:
                base_val = 0.0
            # 保持 value 为 base（浮点） + add_int（整数），不要强制转为 int
            try:
                self.df.at[idx, value_col] = float(base_val) + float(add_int)
            except Exception:
                self.df.at[idx, value_col] = base_val + add_int
        self._compute_base_and_add()

    def save(self, out_path: str):
        if self.df is None:
            raise RuntimeError("数据未加载")
        # 若没有原始文件路径，退回 pandas 全表保存（写入第二行作为 header）
        if not self.path:
            try:
                self.df.to_excel(out_path, index=False, engine="openpyxl")
            except Exception as e:
                raise RuntimeError(f"保存 Excel 失败: {e}")
            return
        try:
            from openpyxl import load_workbook
        except Exception:
            try:
                self.df.to_excel(out_path, index=False, engine="openpyxl")
            except Exception as e2:
                raise RuntimeError(f"保存 Excel 失败（openpyxl 未安装且 pandas 导出失败）: {e2}")
            return

        wb = load_workbook(self.path)
        sheet_name = self.sheet_name if self.sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        header_row = 2  # header=1 对应工作表第2行

        # 建立表头名 -> 列字母 映射（采用工作表现有表头，strip 处理）
        header_map = {}
        for cell in ws[header_row]:
            if cell.value is None:
                continue
            header_map[str(cell.value).strip()] = cell.column_letter

        # 定位 ID 列字母
        id_col_letter = header_map.get("ID")
        if not id_col_letter:
            for k in header_map:
                if str(k).strip().lower() == "id":
                    id_col_letter = header_map[k]
                    break
        if not id_col_letter:
            raise RuntimeError("工作表中未找到 ID 列，无法定位行以进行部分更新。")

        # 构建 excel id -> 行号 映射
        excel_id_map = {}
        for r in range(header_row + 1, ws.max_row + 1):
            cell = ws[f"{id_col_letter}{r}"]
            val = cell.value
            if val is None:
                continue
            excel_id_map[str(val)] = r

        # 对于每一行（按 df），若能定位到工作表行，就把工作表中存在的 header 列全部更新为 df 对应值
        for idx, row in self.df.iterrows():
            id_val = row.get("ID")
            if id_val is None:
                continue
            excel_row = excel_id_map.get(str(id_val))
            if excel_row is None:
                # 如果 ID 行在原表找不到，跳过（避免新增乱位）
                continue
            for hname, col_letter in header_map.items():
                # 仅当 df 中存在该列才写回（这样不会破坏工作表中额外的列）
                if hname not in self.df.columns:
                    continue
                excel_cell = ws[f"{col_letter}{excel_row}"]
                df_val = row.get(hname)
                try:
                    if pd.isna(df_val):
                        excel_cell.value = None
                    else:
                        if isinstance(df_val, int):
                            excel_cell.value = int(df_val)
                        elif isinstance(df_val, float):
                            excel_cell.value = float(df_val)
                        else:
                            excel_cell.value = df_val
                except Exception:
                    excel_cell.value = df_val

        try:
            wb.save(out_path)
        except Exception as e:
            raise RuntimeError(f"保存 Excel 失败: {e}")

    def export_full_red(self, out_path: str):
        if self.df is None:
            raise RuntimeError("数据未加载")
        if not self.path:
            raise RuntimeError("需要原始文件路径才能做部分插入保存")

        try:
            from openpyxl import load_workbook
        except Exception:
            raise RuntimeError("export_full_red 需要 openpyxl，可通过 pip install openpyxl 安装")

        wb = load_workbook(self.path)
        sheet_name = self.sheet_name if self.sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        header_row = 2

        # 读取表头顺序与列字母映射
        headers_in_order = []
        header_map = {}
        for cell in ws[header_row]:
            name = None if cell.value is None else str(cell.value).strip()
            headers_in_order.append(name)
            if name is not None:
                header_map[name] = cell.column_letter

        # 定位 ID 列字母
        id_col_letter = header_map.get("ID")
        if not id_col_letter:
            for k in header_map:
                if k and str(k).strip().lower() == "id":
                    id_col_letter = header_map[k]
                    break
        if not id_col_letter:
            raise RuntimeError("未找到 ID 列，无法执行满红导出")

        # 缓存 worksheet 中每行原始单元格值（按 ID -> row index）
        excel_id_to_row = {}
        for r in range(header_row + 1, ws.max_row + 1):
            cell = ws[f"{id_col_letter}{r}"]
            if cell.value is None:
                continue
            excel_id_to_row[str(cell.value)] = r

        # ATTR 对应列预计算
        attr_cols = {a: self._col_names_for(a) for a in ATTRS}

        # 准备要追加的行数据（按原 df 顺序）
        rows_to_append = []
        for idx, df_row in self.df.iterrows():
            orig_id = df_row.get("ID")
            if orig_id is None:
                continue
            try:
                if isinstance(orig_id, (int, float)):
                    orig_id_str = str(int(float(orig_id)))
                elif isinstance(orig_id, str) and orig_id.strip().isdigit():
                    orig_id_str = str(int(orig_id.strip()))
                else:
                    orig_id_str = str(orig_id)
            except Exception:
                orig_id_str = str(orig_id)
            new_id = None
            if len(orig_id_str) == 3:
                new_id = "50" + orig_id_str
            elif len(orig_id_str) == 2:
                new_id = "500" + orig_id_str
            else:
                continue  # 非 2/3 位跳过

            # 尽量读取工作表中该行的原始单元格值（保留浮点格式）
            excel_row = excel_id_to_row.get(orig_id_str)
            orig_values = {}
            if excel_row:
                for hname, col_letter in header_map.items():
                    orig_values[hname] = ws[f"{col_letter}{excel_row}"].value

            # 构建新行字典（header名 -> 值）
            new_row_vals = {}
            for hname in headers_in_order:
                if hname is None:
                    continue
                # ID 列
                if hname == "ID":
                    new_row_vals[hname] = new_id
                    continue
                # name 列
                if hname == "name":
                    orig_name = orig_values.get("name") if orig_values.get("name") is not None else df_row.get("name", "")
                    new_row_vals[hname] = f"{orig_name}(满红)"
                    continue

                # add_* 列：翻倍并以整数写入（若表中存在该列）
                is_add_col = False
                for a in ATTRS:
                    add_col_name = f"add_{a}"
                    if hname == add_col_name:
                        old_add = df_row.get(add_col_name, 0)
                        try:
                            old_add_f = float(old_add)
                        except Exception:
                            old_add_f = 0.0
                        new_add = int(round(old_add_f * 2))
                        new_row_vals[hname] = new_add
                        is_add_col = True
                        break
                if is_add_col:
                    continue

                # value 列：写为 base + new_add，保留浮点（不强制为整数）
                handled = False
                for a in ATTRS:
                    value_col, init_col, growth_col = attr_cols[a]
                    if hname == value_col:
                        # 计算 base（优先使用 df 中计算好的 base_*，否则用 init/growth）
                        base_val = None
                        if f"base_{a}" in self.df.columns:
                            base_val = df_row.get(f"base_{a}", None)
                        if base_val is None:
                            try:
                                init_v = float(df_row.get(init_col, 0.0)) if init_col in df_row.index else float(orig_values.get(init_col, 0.0) or 0.0)
                            except Exception:
                                init_v = 0.0
                            try:
                                growth_v = float(df_row.get(growth_col, 0.0)) if growth_col in df_row.index else float(orig_values.get(growth_col, 0.0) or 0.0)
                            except Exception:
                                growth_v = 0.0
                            base_val = init_v + growth_v * GROWTH_MULT
                        # old add
                        old_add = df_row.get(f"add_{a}", None)
                        if old_add is None:
                            old_add = orig_values.get(f"add_{a}", 0)
                        try:
                            old_add_f = float(old_add)
                        except Exception:
                            old_add_f = 0.0
                        new_add = int(round(old_add_f * 2))
                        # 新的实际值为 base + new_add（保持浮点，不转为 int）
                        try:
                            new_value = float(base_val) + float(new_add)
                        except Exception:
                            new_value = base_val
                        new_row_vals[hname] = new_value
                        handled = True
                        break
                if handled:
                    continue

                # 其它列：优先使用 worksheet 的原始单元格值，其次使用 df 中的值（不改变类型）
                val = orig_values.get(hname, None)
                if val is None:
                    val = df_row.get(hname, None)
                new_row_vals[hname] = val

            rows_to_append.append(new_row_vals)

        # 将准备好的新行依次追加到表末尾（保持 header 行在第 header_row）
        append_at = ws.max_row + 1
        for new_row in rows_to_append:
            for hname, col_letter in header_map.items():
                cell_ref = f"{col_letter}{append_at}"
                val = new_row.get(hname)
                # add_* 写整数，其它数值保持原类型（float 等）
                ws[cell_ref].value = val
            append_at += 1

        # 保存为 out_path
        try:
            wb.save(out_path)
        except Exception as e:
            raise RuntimeError(f"满红导出保存失败: {e}")

# ---------------- Qt Model & Dialog ----------------
class DataFrameModel(QAbstractTableModel):
    def __init__(self, df: pd.DataFrame, columns: list):
        super().__init__()
        self._df = df
        self._cols = columns

    def rowCount(self, parent=QModelIndex()):
        return 0 if self._df is None else len(self._df)

    def columnCount(self, parent=QModelIndex()):
        return len(self._cols)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or self._df is None:
            return QVariant()
        row = index.row()
        col = self._cols[index.column()]
        if role == Qt.DisplayRole:
            try:
                val = self._df.iloc[row, self._df.columns.get_loc(col)]
            except Exception:
                val = ""
            try:
                if val is None:
                    return ""
                if isinstance(val, float):
                    if abs(val - round(val)) < 1e-9:
                        return str(int(round(val)))
                    return str(val)
                if isinstance(val, (int,)):
                    return str(int(val))
                f = float(val)
                if abs(f - round(f)) < 1e-9:
                    return str(int(round(f)))
                return str(val)
            except Exception:
                return str(val)

        if role == Qt.BackgroundRole:
            try:
                # add_sum 列始终为浅灰
                if col == "add_sum":
                    return QBrush(QColor(240, 240, 240))
                # 若是 add_* 列，且该行的默认加点属性与此列匹配，则浅蓝
                if col.startswith("add_"):
                    try:
                        default_attr = None
                        if "_default_target_attr" in self._df.columns:
                            default_attr = self._df.iloc[row, self._df.columns.get_loc("_default_target_attr")]
                        if default_attr and isinstance(default_attr, str):
                            target_add_col = f"add_{default_attr}"
                            if col == target_add_col:
                                return QBrush(QColor(200, 230, 255))  # 浅蓝
                    except Exception:
                        pass
                # 最后：如果加点和存在且不等于默认值，整行标红（浅红）
                if "add_sum" in self._df.columns:
                    try:
                        add_sum_val = int(self._df.iloc[row, self._df.columns.get_loc("add_sum")])
                        if add_sum_val != int(DEFAULT_ADD_VALUE):
                            return QBrush(QColor(255, 200, 200))
                    except Exception:
                        pass
            except Exception:
                pass
            return QVariant()

        return QVariant()

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return QVariant()
        if orientation == Qt.Horizontal:
            name = self._cols[section]
            # 把 add_sum 显示为“加点和”
            if name == "add_sum":
                return "加点和"
            return name
        return str(section + 1)

class AdjustDialog(QDialog):
    def __init__(self, handler: ExcelHandler, id_or_name: Union[int, str], parent=None):
        super().__init__(parent)
        self.handler = handler
        self.id_or_name = id_or_name
        self.setWindowTitle("调整加点")
        self.setModal(True)
        layout = QVBoxLayout(self)
        hero = handler.get_hero(id_or_name)
        if hero is None:
            layout.addWidget(QLabel("未找到英雄"))
            return
        self.spinboxes = {}
        form = QFormLayout()
        for a in ATTRS:
            sb = QSpinBox()
            sb.setRange(-99999, 99999)
            sb.setSingleStep(1)
            cur = int(round(float(hero.get(f"add_{a}", 0))))
            sb.setValue(cur)
            form.addRow(f"{a} 加点:", sb)
            self.spinboxes[a] = sb
        layout.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.on_ok)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def on_ok(self):
        new_adds = {a: self.spinboxes[a].value() for a in self.spinboxes}
        try:
            self.handler.update_add_points(self.id_or_name, new_adds)
        except Exception as e:
            QMessageBox.critical(self, "更新失败", str(e))
            return
        self.accept()

class SingleAttrEditDialog(QDialog):
    def __init__(self, title: str, current_value: int, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.sb = QSpinBox()
        self.sb.setRange(-99999, 99999)
        self.sb.setValue(int(round(float(current_value or 0))))
        form.addRow("加点值:", self.sb)
        layout.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

# ---------------- Main Window ----------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("属性加点工具")
        self.resize(1000, 640)
        self.handler = ExcelHandler()
        self.current_df: Optional[pd.DataFrame] = None
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        top = QHBoxLayout()
        self.open_btn = QPushButton("打开 Excel")
        self.open_btn.clicked.connect(self.open_file)
        self.save_btn = QPushButton("另存为...")
        self.save_btn.clicked.connect(self.save_file)
        self.save_current_btn = QPushButton("保存")            
        self.save_current_btn.clicked.connect(self.save_current)
        self.full_red_btn = QPushButton("满红导出")
        self.full_red_btn.clicked.connect(self.full_red_export)
        self.show_all_cb = QCheckBox("显示所有英雄")
        self.show_all_cb.stateChanged.connect(self.refresh_table)
        top.addWidget(self.open_btn)
        top.addWidget(self.save_current_btn)                   
        top.addWidget(self.save_btn)
        top.addWidget(self.full_red_btn)
        top.addWidget(self.show_all_cb)
        top.addStretch()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("按 ID 或 名称 搜索并回车")
        self.search_input.returnPressed.connect(self.on_search)
        top.addWidget(self.search_input)
        layout.addLayout(top)
        body = QHBoxLayout()
        self.table = QTableView()
        self.table.doubleClicked.connect(self.on_double_click)
        body.addWidget(self.table, 1)
        layout.addLayout(body)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if not urls:
            return
        path = urls[0].toLocalFile()
        if path:
            self.load_path(path)

    def open_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "打开 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if path:
            self.load_path(path)

    def load_path(self, path: str):
        try:
            self.handler.load(path, sheet_name="hero")
        except Exception as e:
            QMessageBox.critical(self, "读取失败", str(e))
            return
        self.refresh_table()

    def refresh_table(self):
        if self.handler.df is None:
            return
        if self.show_all_cb.isChecked():
            df_show = self.handler.df.copy()
        else:
            df_show = self.handler.df[~self.handler.df["is_default_add"]].copy()
        df_show = df_show.reset_index(drop=False)
        # 原列顺序：ID name is_default_add add_武力 ... add_速度，然后加点和
        add_cols = [f"add_{a}" for a in ATTRS]
        cols = ["ID", "name", "is_default_add"] + [c for c in add_cols if c in df_show.columns]
        # 在 add_速度 后面插入 add_sum（显示名称为 加点和）
        if "add_sum" in df_show.columns:
            cols.append("add_sum")
        cols = [c for c in cols if c in df_show.columns]
        self.current_df = df_show
        model = DataFrameModel(df_show, cols)
        self.table.setModel(model)
        self.table.resizeColumnsToContents()

    def on_search(self):
        term = self.search_input.text().strip()
        if not term:
            self.refresh_table()
            return
        if self.handler.df is None:
            return
        res = self.handler.search(term)
        if res.empty:
            QMessageBox.information(self, "未找到", "没有匹配的英雄")
            return
        self.current_df = res.reset_index(drop=False)
        cols = ["ID", "name", "is_default_add"] + [f"add_{a}" for a in ATTRS]
        cols = [c for c in cols if c in self.current_df.columns]
        self.table.setModel(DataFrameModel(self.current_df, cols))
        self.table.resizeColumnsToContents()

    def on_double_click(self, index: QModelIndex):
        """
        双击处理：
        - 若双击列为 add_* 则弹出 SingleAttrEditDialog 修改该属性的加点值（只应用到内存并刷新）
        - 否则保持原有行为，弹出 AdjustDialog（修改全部属性）
        """
        if self.current_df is None or not index.isValid():
            return
        row = index.row()
        col = index.column()

        # 当前展示列名由模型 cols 列表确定
        model = self.table.model()
        try:
            col_name = model._cols[col]
        except Exception:
            col_name = None

        # 找到原始 handler.df 的索引
        orig_idx = None
        try:
            if "index" in self.current_df.columns:
                orig_idx = int(self.current_df.at[row, "index"])
            else:
                # 使用 ID 定位
                id_val = self.current_df.at[row, "ID"]
                mask = self.handler.df["ID"].astype(str) == str(id_val)
                if mask.any():
                    orig_idx = self.handler.df[mask].index[0]
        except Exception:
            orig_idx = None

        if col_name and col_name.startswith("add_"):
            # 单属性编辑
            cur_val = 0
            try:
                cur_val = int(round(float(self.handler.df.at[orig_idx, col_name])))
            except Exception:
                cur_val = 0
            dlg = SingleAttrEditDialog(col_name, cur_val, parent=self)
            if dlg.exec_():
                new_val = int(dlg.sb.value())
                # 写回 handler.df 仅该 add_ 列，并更新对应 value 列为 base+add（保持浮点）
                try:
                    self.handler.df.at[orig_idx, col_name] = int(new_val)
                    # 同步对应 value 列
                    attr = col_name.replace("add_", "")
                    value_col, init_col, growth_col = self.handler._col_names_for(attr)
                    # 确保 base 存在
                    if f"base_{attr}" not in self.handler.df.columns:
                        init_v = float(self.handler.df.at[orig_idx, init_col]) if init_col in self.handler.df.columns else 0.0
                        growth_v = float(self.handler.df.at[orig_idx, growth_col]) if growth_col in self.handler.df.columns else 0.0
                        self.handler.df.at[orig_idx, f"base_{attr}"] = init_v + growth_v * GROWTH_MULT
                    try:
                        base_val = float(self.handler.df.at[orig_idx, f"base_{attr}"])
                    except Exception:
                        base_val = 0.0
                    # 保持 value 为浮点 base + 整数 add
                    self.handler.df.at[orig_idx, value_col] = float(base_val) + float(new_val)
                except Exception as e:
                    QMessageBox.critical(self, "更新失败", str(e))
                # 重新计算并刷新界面
                try:
                    self.handler._compute_base_and_add()
                except Exception:
                    pass
                self.refresh_table()
            return

        # 不是单属性列，则回退为原先的整体调整弹窗
        if "index" in self.current_df.columns:
            try:
                orig_idx = int(self.current_df.at[row, "index"])
                id_val = self.handler.df.at[orig_idx, "ID"]
            except Exception:
                id_val = self.current_df.at[row, "ID"]
        else:
            id_val = self.current_df.at[row, "ID"]

        dlg = AdjustDialog(self.handler, id_val, parent=self)
        if dlg.exec_():
            self.handler._compute_base_and_add()
            self.refresh_table()

    def save_file(self):
        if self.handler.df is None:
            QMessageBox.information(self, "提示", "当前没有可保存的数据")
            return
        path, _ = QFileDialog.getSaveFileName(self, "保存为", "modified.xlsx", "Excel 文件 (*.xlsx)")
        if not path:
            return
        try:
            self.handler.save(path)
            QMessageBox.information(self, "保存成功", f"已保存到：{path}")
        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e))

    def save_current(self):
        """直接保存当前修改，优先覆盖原始文件，否则弹出保存对话框"""
        if self.handler.df is None:
            QMessageBox.information(self, "提示", "当前没有可保存的数据")
            return
        if self.handler.path:
            try:
                # 尝试保存到原路径（handler.save 会处理部分写回）
                self.handler.save(self.handler.path)
                QMessageBox.information(self, "保存成功", f"已保存到原文件：{self.handler.path}")
            except Exception as e:
                QMessageBox.critical(self, "保存失败", str(e))
        else:
            # 没有原始路径则提示另存为
            path, _ = QFileDialog.getSaveFileName(self, "保存为", "modified.xlsx", "Excel 文件 (*.xlsx)")
            if not path:
                return
            try:
                self.handler.save(path)
                QMessageBox.information(self, "保存成功", f"已保存到：{path}")
            except Exception as e:
                QMessageBox.critical(self, "保存失败", str(e))

    def full_red_export(self):
        if self.handler.df is None:
            QMessageBox.information(self, "提示", "请先打开并加载源 Excel 文件")
            return
        path, _ = QFileDialog.getSaveFileName(self, "满红导出保存为", "full_red.xlsx", "Excel 文件 (*.xlsx)")
        if not path:
            return
        try:
            self.handler.export_full_red(path)
            QMessageBox.information(self, "完成", f"满红导出已保存到：{path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

# ---------------- 运行 ----------------
def main():
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()